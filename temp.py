# # THIS FILE SERVES NO PURPOSE IN THE ACTUAL CODE, ONLY USED FOR TESTING.

from datetime import datetime, timedelta, date
from openpyxl import Workbook
import time
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import (NumericAxis, TextAxis, 
SeriesAxis, DateAxis, ChartLines, NestedInteger)

time.sleep(1)
wb = Workbook()
ws = wb.active

rows = [
    ['Date', 'Batch 1', 'Batch 2', 'Batch 3'],
    [datetime.now().strftime("%m/%d/%Y-%H:%M:%S"), 40, 30, 25],
    [(datetime.now()+timedelta(seconds=1)).strftime("%m/%d/%Y-%H:%M:%S"), 40, 25, 30],
    [(datetime.now()+timedelta(seconds=2)).strftime("%m/%d/%Y-%H:%M:%S"), 50, 30, 45],
    [(datetime.now()+timedelta(seconds=3)).strftime("%m/%d/%Y-%H:%M:%S"), 30, 25, 40],
    [(datetime.now()+timedelta(seconds=4)).strftime("%m/%d/%Y-%H:%M:%S"), 25, 35, 30],
    [(datetime.now()+timedelta(seconds=5)).strftime("%m/%d/%Y-%H:%M:%S"), 20, 40, 35],
]
for row in rows:
    ws.append(row)

data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=ws.max_row)

c2 = LineChart()
c2.title = "Date Axis"
c2.style = 13
c2.y_axis.title = "Size"
c2.y_axis.crossAx = 500 # this is mandatory
c2.x_axis = DateAxis(crossAx=100) # this ia mandatory
c2.x_axis.number_format = 'dd/mm/yy-HH:MM'
c2.x_axis.majorTimeUnit = "days"
c2.x_axis.title = "Date"
c2.x_axis.scaling.min = 2
c2.x_axis.scaling.max = 10
c2.x_axis.minorGridlines = ChartLines()
c2.x_axis.majorTickMark = 'cross'

c2.add_data(data, titles_from_data=True)

# in below we tell where the dates are, in our example dates are in 
# 1st column
dates = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
c2.set_categories(dates)

ws.add_chart(c2, "F1")
wb.save("line.xlsx")

