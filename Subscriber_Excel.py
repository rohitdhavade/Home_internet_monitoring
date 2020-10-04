from SubscriberInterface import SubscriberInterface
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference, Series
from datetime import date, timedelta, datetime
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.axis import NumericAxis, TextAxis, SeriesAxis, DateAxis, ChartLines

wb = Workbook()
ws = wb.create_sheet(title="Report")
dateandtime = datetime.now().strftime("%d.%m.%Y-%H.%M.%S")
dest_filename = f'Report-{dateandtime}.xlsx'
ws.append(['Date Time','Reachability Percent for selected interval'])
x_ax_date = "%d/%m-%H:%M" #specifies format for date in graph

class Subscriber_Excel(SubscriberInterface):
    
    def __init__(self, uptime_dict:dict):
        self.uptime_dict = uptime_dict
        self.points_added:int = 0
        self.minrow:int = 1
        self.anchor_point:int = 1
        self.chart_height:int = 8
        self.chart_width:int = 25

    def update(self, publisher):
        self.uptime_dict = publisher.get_state()

    def update_excel(self):
        latest_key = sorted(self.uptime_dict.keys())[-1]
        print('dictionary is ' + str(self.uptime_dict))
        print('and the latest_key extracted is ' + latest_key.strftime(x_ax_date))
        try:
            print('Append executing')
            ws.append([latest_key.strftime(x_ax_date),self.uptime_dict[latest_key]])
            print('Append Executed')
            #wb.save(filename=dest_filename)
        except PermissionError:
            print('We didn\'t get permission to write to excel file, may it is open somewhere.')
    def update_excel_graph(self):
        self.update_excel()
        x_axis_max_scale = 70
        if self.points_added == x_axis_max_scale:
            self.minrow = self.minrow + self.points_added
            self.points_added = 0
            self.anchor_point = self.anchor_point + 2*self.chart_height + 4
        values = Reference(ws, min_col=2, min_row="%s" %self.minrow, max_row=ws.max_row)
        chart = LineChart()
        chart.height=self.chart_height
        chart.width=self.chart_width
        chart.style = 13
        chart.y_axis.crossAx = 500
        chart.y_axis.title = 'Avg Uptime'
        chart.x_axis = DateAxis(crossAx=100)
        chart.x_axis.title = 'Date'
        chart.x_axis.scaling.max = x_axis_max_scale
        chart.x_axis.scaling.min = 1
        chart.x_axis.majorTickMark = 'cross' 
        # as per our scaling.max value the tickers will be shown on x-axis
        # if scaling.max = 70, then 70 tickers will be shown on x-axis.

        chart.x_axis.majorGridlines = ChartLines()
        chart.y_axis.majorGridlines = ChartLines()

        chart.add_data(values, titles_from_data=True)
        dates = Reference(ws, min_col=1, min_row="%s" %self.minrow, max_row=ws.max_row)
        chart.set_categories(dates) # dates set a x-axis

        try:
            ws.add_chart(chart, "E%s" %self.anchor_point) # e.g.'E1' is anchor pint where graph starts in sheet
            wb.save(filename=dest_filename)
            self.points_added += 1
        except PermissionError:
            print('We didn\'t get permission to write to excel file, may it is open somewhere.')
